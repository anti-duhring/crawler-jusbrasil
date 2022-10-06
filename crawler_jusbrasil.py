from bs4 import BeautifulSoup as beauty
import requests
import cloudscraper
from tkinter.scrolledtext import ScrolledText
from colorama import Fore, Back, Style
import threading
import os
import shutil
import openpyxl

class CrawlerJusbrasil:

    def __init__(self, terms, max_pages, output_path):
        self.output_path = output_path
        self.output_prompt = print
        self.page_content = ''
        self.current_page = 1
        self.max_pages = max_pages
        self.terms = '+'.join(terms.replace(',','').split(' '))
        self.url = f'https://www.jusbrasil.com.br/jurisprudencia/busca?q={self.terms}&p={self.current_page}&idtopico=T10000404&jurisType=SENTENCA'

        try:
            clean_output = threading.Thread(target=self.clean_output)
            clean_output.start()
        except:
            self.myprint('Erro ai configurar planilha de output', 'red')

        try:
            get_page_content = threading.Thread(target=self.get_page_content, args=(self.url,))
            get_page_content.start()
        except:
            self.myprint('Não foi possível fazer request', 'red')

    def get_page_content(self, url):
        scraper = cloudscraper.create_scraper(delay=10, browser='chrome') 
        result = scraper.get(url)
        self.page_content = beauty(result.text, "html.parser")

        # Get the content from this page
        try:
            self.get_search_result()
        except Exception as e:
            print(e)
            self.myprint('Não foi encontrado nenhuma sentença com os termos da pesquisa', 'yellow')
        
        # Loop this function until all pages in self.max_pages are done
        try:
            self.iterate_each_page()
        except Exception as e:
            print(e)
            self.myprint(f"Erro ao acessar página {self.current_page}", 'red')
    
    def get_search_result(self):
        search_result_container = self.page_content.select('.SearchResults-documents')[0]
        
        try:
            self.iterate_each_search_result(search_result=search_result_container)
        except Exception as e:
            self.myprint('Erro ao fazer loop por cada resultado da pesquisa', 'red')
    
    def iterate_each_search_result(self, search_result):    
        for index, result in enumerate(search_result):
            if result.h2 == None:
                return 

            title = result.h2
            result_url = title.a['href']
            content = result.select('.BaseSnippetWrapper-body')
            current_index = index + 1 + (10 * (self.current_page - 1))

            search_data = [content[0].text, '', '', '']

            try:
                self.get_result_data_from_url(url=result_url, search_data=search_data, index=current_index)
            except Exception as e:
                print(e)
                self.myprint(f'Erro ao obter dados da sentença no link {result_url}', 'red')

    def get_result_data_from_url(self, url, search_data, index):
        scraper = cloudscraper.create_scraper(delay=10, browser='chrome') 
        result = scraper.get(url)
        result_content = beauty(result.text, "html.parser")

        title = result_content.select_one('.unprintable').h1
        content = result_content.select_one('article.DocumentPage-content')

        judge_name = content.get_text().split('\n')[0]

        print(judge_name)

        # Data
        search_data[1] = ''
        # Juiz
        search_data[2] = ''
        # Inteiro teor
        search_data[3] = content.text

        try:
            self.append_search_on_output(search_data=search_data)
            self.myprint(f'Sentença {index} adicionada à planilha', 'green')
        except Exception as e:
            print(e)
            self.myprint(f'Erro ao adicionar a sentença {index} na planilha', 'red')

    def iterate_each_page(self):

        self.current_page = self.current_page + 1

        if self.current_page <= self.max_pages:
            new_url = f'https://www.jusbrasil.com.br/jurisprudencia/busca?q={self.terms}&p={self.current_page}&idtopico=T10000404&jurisType=SENTENCA'
            
            get_page_content = threading.Thread(target=self.get_page_content, args=(new_url,))
            get_page_content.start()
        else:
            try:
                self.finish_scrapy()
            except:
                self.myprint('Erro ao finalizar scrapy', 'red')

    def append_search_on_output(self, search_data):
        example_copy_path = f'{self.output_path}/sentenças_pesquisa.xlsx'

        wb = openpyxl.load_workbook(filename=example_copy_path)
        sheet = wb['Sheet1']
        sheet.append(search_data)
        wb.save(example_copy_path)

    def finish_scrapy(self):
        os.startfile(self.output_path)

    def clean_output(self):
        output_filename = os.path.join(os.getcwd(),'outputbot', 'sentenças_pesquisa.xlsx')
        wrkbk_output = openpyxl.load_workbook(filename=output_filename)
        sheet_output = wrkbk_output.active
        sheet_output.delete_rows(2, sheet_output.max_row+1)
        wrkbk_output.save(output_filename)

        try:
            self.copy_output()
        except:
            self.myprint('Erro ao copiar planilha de output', 'red')

    def copy_output(self):
        example_path = os.path.join(os.getcwd(),'outputbot', 'sentenças_pesquisa.xlsx')
        example_copy_path = f'{self.output_path}/sentenças_pesquisa.xlsx'
        shutil.copyfile(example_path, example_copy_path)

    def myprint(self, text, color):
        if self.output_prompt==print:
            colorText = Fore.WHITE
            if color == 'red':
                colorText = Fore.RED
            elif color == 'blue':
                colorText = Fore.BLUE
            elif color == 'green':
                colorText = Fore.GREEN
            elif color == 'yellow':
                colorText = Fore.YELLOW

            print(colorText + text)
            print(Style.RESET_ALL)
        elif isinstance(self.output_prompt, ScrolledText):
            self.output_prompt.tag_config(color, foreground=color)

            self.output_prompt.insert("1.0", text+'\n', color)
            self.output_prompt.update_idletasks()
        else:
            print("myprint - ERROR: {0}".format(str(self.output_prompt)))

terms = 'danos morais corte de energia'
max_pages = 2
output_path = 'C:/Users/mateu/Downloads'
CrawlerJusbrasil(terms=terms, max_pages=max_pages, output_path=output_path)

